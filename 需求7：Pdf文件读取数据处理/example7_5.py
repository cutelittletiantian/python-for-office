import os
import pdfplumber
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet

# 设置工作区
os.chdir(path="./resources/ex7_5/")
# 照片路径
photoListPath = "学生照片"
# 来访学生统计表路径
studentPath = "来访学生统计表.pdf"
# 保存路径
targetPath = "来访学生统计表.xlsx"

# 提取pdf文件内容
studentFile = pdfplumber.open(path_or_fp=studentPath)
# 取出表页
page = studentFile.pages[0]
# 获取表格，跳过表头
table = page.extract_tables()[0]

# 表格写入excel
studentBook = openpyxl.Workbook()  # type: Workbook
studentSheet = studentBook["Sheet"]  # type: Worksheet
for rowData in table:
    studentSheet.append(iterable=rowData)

# 折腾一遭后，点一下保存
studentBook.save(filename=targetPath)

# 获取照片文件列表
photoList = os.listdir(path=photoListPath)
# 遍历表格，跳过表头，加照片
for rowIndex, rowData in enumerate(studentSheet.iter_rows(min_row=2), start=1):
    # 提取姓名
    name = rowData[0].value
    # 加照片
    if f"{name}.png" in photoList:
        # 加载照片文件
        photoPath = Image(img=os.path.join(photoListPath, f"{name}.png"))
        # 加入到行末单元格
        studentSheet.add_image(img=photoPath, anchor=f"D{rowIndex}")
# 再给工作表第4列的表头增加“照片”字样
studentSheet.cell(row=1, column=4).value = "照片"

# 点保存
studentBook.save(filename=targetPath)