import openpyxl
import os

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

orderPath = "./resources/ex4_3/Michael/doc/"

# 表头
sheetHead = ["订单号", "商品ID", "商品名称", "品牌", "类别", "规格", "单价", "数量", "总价", "下单时间"]

allOrderItems = os.listdir(orderPath)
for orderItem in allOrderItems:
    # 完整路径
    itemPath = os.path.join(orderPath, orderItem)

    # 获取工作簿
    orderBook: Workbook = openpyxl.load_workbook(filename=itemPath, data_only=True)
    # 获取工作表
    orderSheet: Worksheet = orderBook["销售订单数据"]

    # 现有商品名
    existProducts = []
    # 遍历行
    for rowData in orderSheet.iter_rows(min_row=2):
        # 获取商品名称
        product = rowData[2].value
        # 添加商品
        if product not in existProducts:
            existProducts.append(product)

    # 创建商品工作表
    for existProduct in existProducts:
        # 建表
        newSheet = orderBook.create_sheet(title=existProduct)
        # 填表头
        newSheet.append(sheetHead)

    # 保存
    orderBook.save(filename=itemPath)
