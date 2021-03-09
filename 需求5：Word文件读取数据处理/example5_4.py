import os
import openpyxl
import docx
from docx.document import Document
from docx.table import Table
from docx.text.run import Run
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# 英语报名表路径
engRegisterPath = "./resources/ex5_4/registration/英语"
# 统计表路径
gatherPath = "./resources/ex5_4/志愿者统计.xlsx"

# 遍历文件
allRegisters = os.listdir(path=engRegisterPath)
# 所有学生的信息
studentInfos = []
# 表转列表，信息提取
for register in allRegisters:
    # 学生信息
    studentInfo = dict()
    # 组装完整路径
    registerPath = os.path.join(engRegisterPath, register)
    # 加载文件
    docReg = docx.Document(docx=registerPath)  # type: Document
    # 加载表格
    table = docReg.tables[0]  # type: Table
    # 读取数据
    studentInfo["姓名"] = table.cell(row_idx=0, col_idx=1).text
    studentInfo["性别"] = table.cell(row_idx=0, col_idx=3).text
    studentInfo["籍贯"] = table.cell(row_idx=0, col_idx=5).text
    studentInfo["学校"] = table.cell(row_idx=1, col_idx=1).text
    studentInfo["年级"] = table.cell(row_idx=1, col_idx=3).text
    studentInfo["专业"] = table.cell(row_idx=1, col_idx=5).text
    studentInfo["联系电话"] = table.cell(row_idx=2, col_idx=1).text
    studentInfo["QQ/微信号"] = table.cell(row_idx=2, col_idx=3).text
    studentInfo["是否愿意服从调剂"] = table.cell(row_idx=2, col_idx=5).text

    # 添加到列表
    studentInfos.append(studentInfo)

# 加载excel表格
engBook = openpyxl.load_workbook(filename=gatherPath)  # type: Workbook
# 加载英语工作表
engSheet = engBook["英语"]  # type: Worksheet
# 遍历列表，写入excel文档
for student in studentInfos:
    studentData = [something for something in student.values()]
    engSheet.append(iterable=studentData)

# 保存
engBook.save(filename=gatherPath)