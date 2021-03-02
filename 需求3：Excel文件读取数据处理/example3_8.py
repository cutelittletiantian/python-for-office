import openpyxl
import os

from openpyxl.worksheet.worksheet import Worksheet

resourcePath = "./resources/ex3_8"
os.chdir(resourcePath)

# 打开并加载表
orderSheet: Worksheet = openpyxl.load_workbook(filename="平台销售订单.xlsx", data_only=True)["销售订单数据"]
sysSheet: Worksheet = openpyxl.load_workbook(filename="系统数据.xlsx", data_only=True)["订单数据"]

# 遍历订单表的行数据，读取出订单id
order_id = []
for rowData in orderSheet.iter_rows(min_row=2, min_col=1, max_col=1):
    order_id.append(rowData[0].value)

# 遍历系统数据，读取
order_id_sys = []
for rowData in sysSheet.iter_rows(min_row=2, min_col=2, max_col=2):
    order_id_sys.append(rowData[0].value)

# 订单号 xxxxxx 不在系统数据中
for data in order_id:
    if data not in order_id_sys:
        print(f"订单号 {data} 不在系统数据中")

# 订单号 xxxxxx 不在商城数据中
for data in order_id_sys:
    if data not in order_id:
        print(f"订单号 {data} 不在商城数据中")