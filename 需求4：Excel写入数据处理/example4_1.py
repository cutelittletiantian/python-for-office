import openpyxl
import os

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

os.chdir("./resources/ex4_1/")

newWb: Workbook = openpyxl.Workbook()

# 默认表改名
courseSheet: Worksheet = newWb["Sheet"]
courseSheet.title = "七年级1班课表"

# 2~10班
for classNum in range(2, 11):
    newWb.create_sheet(title=f"七年级{classNum}班课表")

for curSheet in newWb.worksheets:
    curSheet.append(["", "星期一", "星期二", "星期三", "星期四", "星期五"])
    curSheet.merge_cells(start_row=2, end_row=4, start_column=1, end_column=1)
    curSheet["A2"] = "上午"
    curSheet.merge_cells(start_row=5, end_row=6, start_column=1, end_column=1)
    curSheet["A5"] = "下午"

newWb.save(filename="七年级课表.xlsx")